"""
Export module for generating PDF and Excel reports from variant data.
"""

import os
import logging
from datetime import datetime
import config

logger = logging.getLogger(__name__)

# Try to import optional dependencies
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    logger.warning("openpyxl not available. Excel export will be disabled.")

try:
    from docx2pdf import convert
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False
    logger.warning("docx2pdf not available. PDF export will be disabled.")


def export_to_excel(data, output_path):
    """
    Export variant data to Excel format, appending to collective files by gene group.
    
    Args:
        data: Dictionary containing variant and patient information
        output_path: Directory path where the Excel file should be saved
        
    Returns:
        str: Path to the generated Excel file, or None if export failed
    """
    if not EXCEL_AVAILABLE:
        logger.error("Excel export not available - openpyxl not installed")
        return None
    
    try:
        logger.info("Starting Excel export...")
        
        if not data:
            logger.error("No data provided for Excel export")
            return None
        
        # Validate output path
        if not output_path or '..' in output_path:
            logger.error(f"Invalid output path: {output_path}")
            return None
        
        # Ensure output directory exists
        os.makedirs(output_path, exist_ok=True)
        
        # Determine category and collective filename
        category = data.get("Category", "Övrigt")
        
        # Map category to gene group for collective file
        if category == "Medfodd anemi" or "anemi" in category.lower():
            collective_filename = "Medfodd_anemi_collective.xlsx"
        elif category == "Koagulation" or "koagulation" in category.lower():
            collective_filename = "Koagulation_collective.xlsx"
        else:
            collective_filename = "Ovrigt_collective.xlsx"
        
        collective_path = os.path.join(output_path, collective_filename)
        
        # Load existing workbook or create new one
        if os.path.exists(collective_path):
            wb = openpyxl.load_workbook(collective_path)
            ws = wb.active
            logger.info(f"Appending to existing file: {collective_path}")
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Variant Data"
            
            # Define styles
            header_font = Font(bold=True, size=12)
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # Add header row
            headers = [
                "LID-NR", "Gene", "Nucleotide Change", "Protein Change", 
                "Zygosity", "Inheritance", "ACMG Assessment", "ClinVar Info",
                "Further Studies", "Transcript", "Disease", "Category",
                "Proband", "Genotype", "Phenotype", "Sequencing Method", "Exon",
                "Timestamp"
            ]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            logger.info(f"Created new collective file: {collective_path}")
        
        # Find next empty row
        row = ws.max_row + 1
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        if data.get("Normalfynd", False):
            # Normal finding
            ws.cell(row=row, column=1).value = data.get("LID-NR", "")
            ws.cell(row=row, column=2).value = data.get("Gene", "")
            ws.cell(row=row, column=3).value = "Normal finding"
            ws.cell(row=row, column=10).value = data.get("Transcript", "")
            ws.cell(row=row, column=11).value = data.get("Disease", "")
            ws.cell(row=row, column=12).value = data.get("Category", "")
            ws.cell(row=row, column=13).value = data.get("Proband", "")
            ws.cell(row=row, column=14).value = data.get("Genotype", "")
            ws.cell(row=row, column=15).value = data.get("Phenotype", "")
            ws.cell(row=row, column=16).value = data.get("Sequencing method", "")
            ws.cell(row=row, column=17).value = data.get("Exon", "")
            ws.cell(row=row, column=18).value = timestamp
        else:
            # Variants
            for variant in data.get("variants", []):
                ws.cell(row=row, column=1).value = data.get("LID-NR", "")
                ws.cell(row=row, column=2).value = variant.get("Gene", "")
                ws.cell(row=row, column=3).value = variant.get("Nucleotide change", "")
                ws.cell(row=row, column=4).value = variant.get("Protein change", "")
                ws.cell(row=row, column=5).value = variant.get("Zygosity", "")
                ws.cell(row=row, column=6).value = variant.get("Inheritance", "")
                ws.cell(row=row, column=7).value = variant.get("ACMG criteria assessment", "")
                ws.cell(row=row, column=8).value = variant.get("ClinVar and hemophilia database reports", "")
                ws.cell(row=row, column=9).value = variant.get("Further studies", "")
                ws.cell(row=row, column=10).value = data.get("Transcript", "")
                ws.cell(row=row, column=11).value = data.get("Disease", "")
                ws.cell(row=row, column=12).value = data.get("Category", "")
                ws.cell(row=row, column=13).value = data.get("Proband", "")
                ws.cell(row=row, column=14).value = data.get("Genotype", "")
                ws.cell(row=row, column=15).value = data.get("Phenotype", "")
                ws.cell(row=row, column=16).value = data.get("Sequencing method", "")
                ws.cell(row=row, column=17).value = data.get("Exon", "")
                ws.cell(row=row, column=18).value = timestamp
                row += 1
        
        # Auto-adjust column widths (only for new files)
        if ws.max_row <= 2:  # Only header and first data row
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save the collective file
        wb.save(collective_path)
        logger.info(f"Excel data appended to collective file: {collective_path}")
        return collective_path
        
    except Exception as e:
        logger.error(f"Error exporting to Excel: {e}", exc_info=True)
        return None


def export_to_pdf(docx_path):
    """
    Convert a Word document to PDF format.
    
    Args:
        docx_path: Path to the Word document to convert
        
    Returns:
        str: Path to the generated PDF file, or None if conversion failed
    """
    if not PDF_AVAILABLE:
        logger.error("PDF export not available - docx2pdf not installed")
        return None
    
    try:
        logger.info(f"Starting PDF conversion for {docx_path}...")
        
        if not os.path.exists(docx_path):
            logger.error(f"Word document not found: {docx_path}")
            return None
        
        # Generate PDF path (same name, different extension)
        pdf_path = os.path.splitext(docx_path)[0] + ".pdf"
        
        # Convert to PDF
        convert(docx_path, pdf_path)
        
        logger.info(f"PDF file created: {pdf_path}")
        return pdf_path
        
    except Exception as e:
        logger.error(f"Error converting to PDF: {e}", exc_info=True)
        return None


def is_excel_available():
    """Check if Excel export is available."""
    return EXCEL_AVAILABLE


def is_pdf_available():
    """Check if PDF export is available."""
    return PDF_AVAILABLE
